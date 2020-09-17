from openpyxl import load_workbook
import sqlite3


def load_details(filename, path):
    con = sqlite3.connect(path)
    cur = con.cursor()
    wb1 = load_workbook(filename=filename)['Лист1']
    data = [i for i in wb1]
    for row in data[1:]:
        cur.execute("""INSERT into details (title, category) values (?, ?)""", (row[1].value, row[2].value))

    con.commit() 


def load_drones(filename, path):
    con = sqlite3.connect(path)
    cur = con.cursor()
    wb1 = load_workbook(filename=filename)['Лист1']
    data = [i for i in wb1]
    for i in data[1:]:
        if str(i[2].value).isdigit():
            cur.execute("""INSERT into drones (title, price) values (?, ?)""", (i[1].value, i[2].value))
        else:
            print(i[1].value, i[2].value, "is wrong")

    con.commit()


def load_tech_cards(filename, path):
    con = sqlite3.connect(path)
    cur = con.cursor()
    wb1 = load_workbook(filename=filename)['Лист1']
    data = [i for i in wb1]
    drone_id = 0
    for i in data[1:]:
        if i[0].value is not None:
            if str(i[3].value).isdigit():
                cur.execute("""INSERT into tech_cards (drone_id, title, amount) values (?, ?, ?)""",
                            (i[0].value, i[2].value, i[3].value))
                drone_id = i[0].value
            else:
                print(i[1].value, i[2].value, "is wrong")
        else:
            if str(i[3].value).isdigit():
                cur.execute("""INSERT into tech_cards (drone_id, title, amount) values (?, ?, ?)""",
                            (drone_id, i[2].value, i[3].value))
            else:
                print(i[1].value, i[2].value, "is wrong")
    con.commit()
