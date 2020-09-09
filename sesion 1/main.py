from openpyxl import load_workbook
import sqlite3


def load_details(filename):
    con = sqlite3.connect("nti_base.db")
    cur = con.cursor()
    wb1 = load_workbook(filename=filename)['Лист1']
    data = [i for i in wb1]
    for i in data[1:]:
        cur.execute("""INSERT into details (title, category) values (?, ?)""", (i[1].value, i[2].value))

    con.commit() 


def load_drones(filename):
    con = sqlite3.connect("nti_base.db")
    cur = con.cursor()
    wb1 = load_workbook(filename=filename)['Лист1']
    data = [i for i in wb1]
    for i in data[1:]:
        print(type(i[2].value))
        if type(i[2].value) == float:
            cur.execute("""INSERT into drones (title, price) values (?, ?)""", (i[1].value, i[2].value))

    con.commit()


def load_tech_cards(filename):
    con = sqlite3.connect("nti_base.db")
    cur = con.cursor()
    wb1 = load_workbook(filename=filename)['Лист1']
    data = [i for i in wb1]
    drone_id = 0
    for i in data[1:]:
        if i[0].value is not None:
            cur.execute("""INSERT into tech_cards (drone_id, title, amount) values (?, ?, ?)""",
                        (i[0].value, i[2].value, i[3].value))
            drone_id = i[0].value
        else:
            cur.execute("""INSERT into tech_cards (drone_id, title, amount) values (?, ?, ?)""",
                        (drone_id, i[2].value, i[3].value))
    con.commit()


load_drones("СписокДронов.xlsx")